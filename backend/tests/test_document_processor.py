"""
Tests for Document Processor Infrastructure

Tests the DocumentProcessor base class, DocumentProcessorFactory,
and ExcelDocumentProcessor adapter.
"""

import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook

from src.services.document_processor import (
    DocumentProcessor,
    DocumentType,
    TextSegment,
    ProcessingResult,
    DocumentProcessorFactory,
    apply_interleaved_mode,
    apply_output_mode,
    apply_auto_append,
)
from src.services.excel_document_processor import ExcelDocumentProcessor


class TestTextSegment:
    """Tests for TextSegment dataclass."""
    
    def test_text_segment_creation(self):
        """Test creating a TextSegment with all fields."""
        segment = TextSegment(
            id="1",
            text="Hello World",
            location="Sheet1!A1",
            metadata={"row": 1, "column": 1}
        )
        
        assert segment.id == "1"
        assert segment.text == "Hello World"
        assert segment.location == "Sheet1!A1"
        assert segment.metadata == {"row": 1, "column": 1}
    
    def test_text_segment_default_metadata(self):
        """Test TextSegment with default empty metadata."""
        segment = TextSegment(
            id="1",
            text="Test",
            location="Page 1"
        )
        
        assert segment.metadata == {}


class TestProcessingResult:
    """Tests for ProcessingResult dataclass."""
    
    def test_processing_result_success(self):
        """Test creating a successful ProcessingResult."""
        result = ProcessingResult(
            success=True,
            segments_total=100,
            segments_translated=100,
            output_path=Path("/output/file.xlsx")
        )
        
        assert result.success is True
        assert result.segments_total == 100
        assert result.segments_translated == 100
        assert result.output_path == Path("/output/file.xlsx")
        assert result.error is None
    
    def test_processing_result_failure(self):
        """Test creating a failed ProcessingResult."""
        result = ProcessingResult(
            success=False,
            segments_total=0,
            segments_translated=0,
            error="File not found"
        )
        
        assert result.success is False
        assert result.error == "File not found"
        assert result.output_path is None


class TestDocumentType:
    """Tests for DocumentType enum."""
    
    def test_document_type_values(self):
        """Test DocumentType enum values."""
        assert DocumentType.EXCEL.value == "excel"
        assert DocumentType.WORD.value == "word"
        assert DocumentType.POWERPOINT.value == "powerpoint"
        assert DocumentType.PDF.value == "pdf"
        assert DocumentType.TEXT.value == "text"
        assert DocumentType.MARKDOWN.value == "markdown"


class TestDocumentProcessorFactory:
    """Tests for DocumentProcessorFactory."""
    
    def test_factory_initialization(self):
        """Test factory initializes with empty registry."""
        factory = DocumentProcessorFactory()
        assert factory.get_supported_extensions() == []
    
    def test_register_processor(self):
        """Test registering a processor."""
        factory = DocumentProcessorFactory()
        processor = ExcelDocumentProcessor()
        
        factory.register(processor)
        
        assert '.xlsx' in factory.get_supported_extensions()
    
    def test_get_processor_for_supported_file(self):
        """Test getting processor for supported file type."""
        factory = DocumentProcessorFactory()
        processor = ExcelDocumentProcessor()
        factory.register(processor)
        
        result = factory.get_processor(Path("test.xlsx"))
        
        assert result is not None
        assert isinstance(result, ExcelDocumentProcessor)
    
    def test_get_processor_for_unsupported_file(self):
        """Test getting processor for unsupported file type."""
        factory = DocumentProcessorFactory()
        processor = ExcelDocumentProcessor()
        factory.register(processor)
        
        result = factory.get_processor(Path("test.txt"))
        
        assert result is None
    
    def test_is_supported(self):
        """Test checking if file type is supported."""
        factory = DocumentProcessorFactory()
        processor = ExcelDocumentProcessor()
        factory.register(processor)
        
        assert factory.is_supported(Path("test.xlsx")) is True
        assert factory.is_supported(Path("test.XLSX")) is True  # Case insensitive
        assert factory.is_supported(Path("test.docx")) is False
    
    def test_get_document_type(self):
        """Test getting document type for a file."""
        factory = DocumentProcessorFactory()
        processor = ExcelDocumentProcessor()
        factory.register(processor)
        
        assert factory.get_document_type(Path("test.xlsx")) == DocumentType.EXCEL
        assert factory.get_document_type(Path("test.docx")) is None


class TestExcelDocumentProcessor:
    """Tests for ExcelDocumentProcessor."""
    
    def test_supported_extensions(self):
        """Test supported extensions."""
        processor = ExcelDocumentProcessor()
        assert processor.supported_extensions == ['.xlsx']
    
    def test_document_type(self):
        """Test document type."""
        processor = ExcelDocumentProcessor()
        assert processor.document_type == DocumentType.EXCEL
    
    def test_generate_output_filename(self):
        """Test output filename generation."""
        processor = ExcelDocumentProcessor()
        
        filename = processor.generate_output_filename(Path("document.xlsx"))
        assert filename == "document_vi.xlsx"
        
        filename = processor.generate_output_filename(Path("report.xlsx"), "en")
        assert filename == "report_en.xlsx"
    
    @pytest.mark.asyncio
    async def test_extract_text_from_excel(self):
        """Test extracting text from an Excel file."""
        processor = ExcelDocumentProcessor()
        
        # Create a test Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws['A1'] = "Hello"
            ws['B1'] = "World"
            ws['A2'] = "Test"
            wb.save(f.name)
            
            file_path = Path(f.name)
        
        try:
            segments = await processor.extract_text(file_path)
            
            assert len(segments) == 3
            texts = [s.text for s in segments]
            assert "Hello" in texts
            assert "World" in texts
            assert "Test" in texts
            
            # Check metadata
            for segment in segments:
                assert "worksheet_name" in segment.metadata
                assert "row" in segment.metadata
                assert "column" in segment.metadata
        finally:
            file_path.unlink()
    
    @pytest.mark.asyncio
    async def test_validate_file_success(self):
        """Test validating a valid Excel file."""
        processor = ExcelDocumentProcessor()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            wb.save(f.name)
            file_path = Path(f.name)
        
        try:
            is_valid, error = await processor.validate_file(file_path)
            assert is_valid is True
            assert error is None
        finally:
            file_path.unlink()
    
    @pytest.mark.asyncio
    async def test_validate_file_not_found(self):
        """Test validating a non-existent file."""
        processor = ExcelDocumentProcessor()
        
        is_valid, error = await processor.validate_file(Path("/nonexistent/file.xlsx"))
        
        assert is_valid is False
        assert "not found" in error.lower()
    
    @pytest.mark.asyncio
    async def test_validate_file_wrong_extension(self):
        """Test validating a file with wrong extension."""
        processor = ExcelDocumentProcessor()
        
        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as f:
            f.write(b"test")
            file_path = Path(f.name)
        
        try:
            is_valid, error = await processor.validate_file(file_path)
            assert is_valid is False
            assert "unsupported" in error.lower()
        finally:
            file_path.unlink()
    
    @pytest.mark.asyncio
    async def test_write_translated(self):
        """Test writing translated text back to Excel."""
        processor = ExcelDocumentProcessor()
        
        # Create a test Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws['A1'] = "Hello"
            ws['B1'] = "World"
            wb.save(f.name)
            input_path = Path(f.name)
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = Path(f.name)
        
        try:
            # Extract text
            segments = await processor.extract_text(input_path)
            
            # Create translations
            translations = ["Xin chào", "Thế giới"]
            
            # Write translated (with auto_append=False to replace original text)
            success = await processor.write_translated(
                input_path, segments, translations, output_path, auto_append=False
            )
            
            assert success is True
            assert output_path.exists()
            
            # Verify the translated content
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active
            assert ws['A1'].value == "Xin chào"
            assert ws['B1'].value == "Thế giới"
        finally:
            input_path.unlink()
            if output_path.exists():
                output_path.unlink()


class TestApplyInterleavedMode:
    """Tests for apply_interleaved_mode function.
    
    Validates Requirements: 4.1, 4.2, 4.3, 4.4, 4.5
    """
    
    def test_single_line_interleaving(self):
        """Test interleaving single line texts.
        
        Validates: Requirements 4.1, 4.2
        """
        original = "Hello"
        translated = "Xin chào"
        
        result = apply_interleaved_mode(original, translated)
        
        assert result == "Hello\nXin chào"
    
    def test_multi_line_equal_count_interleaving(self):
        """Test interleaving multi-line texts with equal line counts.
        
        Validates: Requirements 4.1, 4.2
        """
        original = "Line 1\nLine 2\nLine 3"
        translated = "Dòng 1\nDòng 2\nDòng 3"
        
        result = apply_interleaved_mode(original, translated)
        
        expected = "Line 1\nDòng 1\nLine 2\nDòng 2\nLine 3\nDòng 3"
        assert result == expected
    
    def test_original_has_more_lines(self):
        """Test interleaving when original has more lines than translated.
        
        Validates: Requirements 4.3
        """
        original = "Line 1\nLine 2\nLine 3\nLine 4"
        translated = "Dòng 1\nDòng 2"
        
        result = apply_interleaved_mode(original, translated)
        
        # Expected: Line 1, Dòng 1, Line 2, Dòng 2, Line 3, Line 4
        expected = "Line 1\nDòng 1\nLine 2\nDòng 2\nLine 3\nLine 4"
        assert result == expected
    
    def test_translated_has_more_lines(self):
        """Test interleaving when translated has more lines than original.
        
        Validates: Requirements 4.4
        """
        original = "Line 1\nLine 2"
        translated = "Dòng 1\nDòng 2\nDòng 3\nDòng 4"
        
        result = apply_interleaved_mode(original, translated)
        
        # Expected: Line 1, Dòng 1, Line 2, Dòng 2, Dòng 3, Dòng 4
        expected = "Line 1\nDòng 1\nLine 2\nDòng 2\nDòng 3\nDòng 4"
        assert result == expected
    
    def test_equal_texts_no_duplication(self):
        """Test that equal texts return original without duplication.
        
        Validates: Requirements 4.5
        """
        text = "Same text\nOn multiple lines"
        
        result = apply_interleaved_mode(text, text)
        
        assert result == text
    
    def test_equal_texts_with_whitespace_no_duplication(self):
        """Test that texts equal after stripping return without duplication.
        
        Validates: Requirements 4.5
        """
        original = "  Same text  "
        translated = "Same text"
        
        result = apply_interleaved_mode(original, translated)
        
        # Should return translated (no duplication)
        assert result == translated
    
    def test_empty_original(self):
        """Test interleaving with empty original text."""
        original = ""
        translated = "Dòng 1\nDòng 2"
        
        result = apply_interleaved_mode(original, translated)
        
        # Empty string splits to [''], so we get: '', Dòng 1, Dòng 2
        expected = "\nDòng 1\nDòng 2"
        assert result == expected
    
    def test_empty_translated(self):
        """Test interleaving with empty translated text."""
        original = "Line 1\nLine 2"
        translated = ""
        
        result = apply_interleaved_mode(original, translated)
        
        # Empty string splits to [''], so we get: Line 1, '', Line 2
        expected = "Line 1\n\nLine 2"
        assert result == expected
    
    def test_both_empty(self):
        """Test interleaving with both texts empty."""
        original = ""
        translated = ""
        
        result = apply_interleaved_mode(original, translated)
        
        # Both equal after strip, so return translated
        assert result == ""
    
    def test_preserves_empty_lines_in_original(self):
        """Test that empty lines in original are preserved."""
        original = "Line 1\n\nLine 3"
        translated = "Dòng 1\nDòng 2\nDòng 3"
        
        result = apply_interleaved_mode(original, translated)
        
        # Expected: Line 1, Dòng 1, '', Dòng 2, Line 3, Dòng 3
        expected = "Line 1\nDòng 1\n\nDòng 2\nLine 3\nDòng 3"
        assert result == expected
    
    def test_preserves_empty_lines_in_translated(self):
        """Test that empty lines in translated are preserved."""
        original = "Line 1\nLine 2\nLine 3"
        translated = "Dòng 1\n\nDòng 3"
        
        result = apply_interleaved_mode(original, translated)
        
        # Expected: Line 1, Dòng 1, Line 2, '', Line 3, Dòng 3
        expected = "Line 1\nDòng 1\nLine 2\n\nLine 3\nDòng 3"
        assert result == expected


class TestApplyOutputMode:
    """Tests for apply_output_mode dispatcher function.
    
    Validates: Requirements 4.6
    """
    
    def test_replace_mode_returns_translated_only(self):
        """Test replace mode (both flags false) returns translated text only.
        
        Validates: Requirements 4.6
        """
        original = "Hello World"
        translated = "Xin chào Thế giới"
        
        result = apply_output_mode(original, translated, auto_append=False, interleaved_mode=False)
        
        assert result == translated
    
    def test_append_mode_appends_translated_after_original(self):
        """Test append mode appends translated text after original.
        
        Validates: Requirements 4.6
        """
        original = "Hello World"
        translated = "Xin chào Thế giới"
        
        result = apply_output_mode(original, translated, auto_append=True, interleaved_mode=False)
        
        expected = f"{original}\n{translated}"
        assert result == expected
    
    def test_interleaved_mode_interleaves_lines(self):
        """Test interleaved mode interleaves original and translated lines.
        
        Validates: Requirements 4.6
        """
        original = "Line 1\nLine 2"
        translated = "Dòng 1\nDòng 2"
        
        result = apply_output_mode(original, translated, auto_append=False, interleaved_mode=True)
        
        expected = "Line 1\nDòng 1\nLine 2\nDòng 2"
        assert result == expected
    
    def test_interleaved_mode_takes_precedence_over_append(self):
        """Test that interleaved mode takes precedence when both flags are true.
        
        Note: This scenario should be prevented by validation, but the function
        handles it by prioritizing interleaved mode.
        
        Validates: Requirements 4.6
        """
        original = "Line 1\nLine 2"
        translated = "Dòng 1\nDòng 2"
        
        # Even if both flags are true, interleaved mode takes precedence
        result = apply_output_mode(original, translated, auto_append=True, interleaved_mode=True)
        
        expected = "Line 1\nDòng 1\nLine 2\nDòng 2"
        assert result == expected
    
    def test_replace_mode_with_multiline_text(self):
        """Test replace mode with multi-line text."""
        original = "Line 1\nLine 2\nLine 3"
        translated = "Dòng 1\nDòng 2\nDòng 3"
        
        result = apply_output_mode(original, translated, auto_append=False, interleaved_mode=False)
        
        assert result == translated
    
    def test_append_mode_with_multiline_text(self):
        """Test append mode with multi-line text."""
        original = "Line 1\nLine 2"
        translated = "Dòng 1\nDòng 2"
        
        result = apply_output_mode(original, translated, auto_append=True, interleaved_mode=False)
        
        expected = f"{original}\n{translated}"
        assert result == expected
    
    def test_replace_mode_with_empty_translated(self):
        """Test replace mode with empty translated text."""
        original = "Hello World"
        translated = ""
        
        result = apply_output_mode(original, translated, auto_append=False, interleaved_mode=False)
        
        assert result == ""
    
    def test_append_mode_with_equal_texts(self):
        """Test append mode when original and translated are equal (no duplication)."""
        text = "Same text"
        
        result = apply_output_mode(text, text, auto_append=True, interleaved_mode=False)
        
        # apply_auto_append should avoid duplication when texts are equal
        assert result == text
    
    def test_interleaved_mode_with_equal_texts(self):
        """Test interleaved mode when original and translated are equal (no duplication)."""
        text = "Same text\nOn multiple lines"
        
        result = apply_output_mode(text, text, auto_append=False, interleaved_mode=True)
        
        # apply_interleaved_mode should avoid duplication when texts are equal
        assert result == text
    
    def test_interleaved_mode_with_unequal_line_counts(self):
        """Test interleaved mode handles unequal line counts correctly."""
        original = "Line 1\nLine 2\nLine 3"
        translated = "Dòng 1"
        
        result = apply_output_mode(original, translated, auto_append=False, interleaved_mode=True)
        
        # Expected: Line 1, Dòng 1, Line 2, Line 3
        expected = "Line 1\nDòng 1\nLine 2\nLine 3"
        assert result == expected
    
    def test_replace_mode_preserves_special_characters(self):
        """Test replace mode preserves special characters in translated text."""
        original = "Hello"
        translated = "Xin chào! 你好 🌍"
        
        result = apply_output_mode(original, translated, auto_append=False, interleaved_mode=False)
        
        assert result == translated
    
    def test_append_mode_preserves_whitespace(self):
        """Test append mode preserves whitespace in both texts."""
        original = "  Hello  "
        translated = "  Xin chào  "
        
        result = apply_output_mode(original, translated, auto_append=True, interleaved_mode=False)
        
        expected = f"{original}\n{translated}"
        assert result == expected
