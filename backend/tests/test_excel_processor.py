"""Tests for async ExcelProcessor."""

import asyncio
import tempfile
from pathlib import Path
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from src.services.excel_processor import ExcelProcessor, CellData, WorksheetProgress


@pytest.fixture
def excel_processor():
    """Create an ExcelProcessor instance for testing."""
    return ExcelProcessor()


@pytest.fixture
def sample_workbook():
    """Create a sample workbook with test data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    
    # Add some test data with formatting
    ws['A1'] = "Hello"
    ws['A1'].font = Font(bold=True, size=14, color="FF0000")
    ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws['B1'] = "World"
    ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws['A2'] = "Test"
    ws['A2'].border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add a formula cell (should be skipped)
    ws['C1'] = "=SUM(A1:B1)"
    
    # Add an empty cell (should be skipped)
    ws['D1'] = None
    
    # Add a whitespace-only cell (should be skipped)
    ws['E1'] = "   "
    
    return wb


@pytest.fixture
def temp_excel_file(sample_workbook):
    """Create a temporary Excel file for testing."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = Path(f.name)
        sample_workbook.save(str(temp_path))
    
    yield temp_path
    
    # Cleanup
    if temp_path.exists():
        temp_path.unlink()


class TestExcelProcessor:
    """Unit tests for ExcelProcessor."""
    
    @pytest.mark.asyncio
    async def test_load_workbook_success(self, excel_processor, temp_excel_file):
        """Test loading a valid Excel workbook."""
        workbook = await excel_processor.load_workbook(temp_excel_file)
        
        assert workbook is not None
        assert len(workbook.worksheets) > 0
        assert workbook.active.title == "Test Sheet"
    
    @pytest.mark.asyncio
    async def test_load_workbook_nonexistent_file(self, excel_processor):
        """Test loading a non-existent file returns None."""
        nonexistent_path = Path("/nonexistent/file.xlsx")
        workbook = await excel_processor.load_workbook(nonexistent_path)
        
        assert workbook is None
    
    @pytest.mark.asyncio
    async def test_iterate_cells_in_worksheet(self, excel_processor, sample_workbook):
        """Test iterating cells in a worksheet."""
        worksheet = sample_workbook.active
        cells = await excel_processor.iterate_cells_in_worksheet(worksheet)
        
        # Should have 3 cells: A1, B1, A2 (C1 is formula, D1 is empty, E1 is whitespace)
        assert len(cells) == 3
        
        # Verify cell data
        cell_values = [cell.value for cell in cells]
        assert "Hello" in cell_values
        assert "World" in cell_values
        assert "Test" in cell_values
    
    @pytest.mark.asyncio
    async def test_iterate_cells_skips_formulas(self, excel_processor, sample_workbook):
        """Test that formula cells are skipped."""
        worksheet = sample_workbook.active
        cells = await excel_processor.iterate_cells_in_worksheet(worksheet)
        
        # Formula cell should not be in the results
        for cell in cells:
            assert not cell.has_formula
            assert not str(cell.value).startswith('=')
    
    @pytest.mark.asyncio
    async def test_update_cell_preserves_formatting(self, excel_processor, sample_workbook):
        """Test that updating a cell preserves its formatting."""
        worksheet = sample_workbook.active
        cell = worksheet['A1']
        
        # Store original formatting
        original_font_bold = cell.font.bold
        original_font_size = cell.font.size
        original_font_color = cell.font.color
        original_fill = cell.fill.start_color
        
        # Update the cell
        await excel_processor.update_cell(cell, "Updated Text")
        
        # Verify value changed
        assert cell.value == "Updated Text"
        
        # Verify formatting preserved
        assert cell.font.bold == original_font_bold
        assert cell.font.size == original_font_size
        assert cell.font.color == original_font_color
        assert cell.fill.start_color == original_fill
    
    @pytest.mark.asyncio
    async def test_save_workbook_success(self, excel_processor, sample_workbook):
        """Test saving a workbook to output directory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / "source.xlsx"
            output_dir = Path(tmpdir) / "output"
            
            # Save the workbook
            output_path = await excel_processor.save_workbook(
                sample_workbook,
                source_path,
                output_dir,
                language_suffix="vi"
            )
            
            assert output_path is not None
            assert output_path.exists()
            assert output_path.name == "source_vi.xlsx"
            assert output_path.parent == output_dir
    
    @pytest.mark.asyncio
    async def test_save_workbook_creates_output_directory(self, excel_processor, sample_workbook):
        """Test that save_workbook creates output directory if it doesn't exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / "source.xlsx"
            output_dir = Path(tmpdir) / "nonexistent" / "output"
            
            # Output directory doesn't exist yet
            assert not output_dir.exists()
            
            # Save the workbook
            output_path = await excel_processor.save_workbook(
                sample_workbook,
                source_path,
                output_dir
            )
            
            assert output_path is not None
            assert output_dir.exists()
            assert output_path.exists()
    
    @pytest.mark.asyncio
    async def test_process_worksheets_concurrently(self, excel_processor, sample_workbook):
        """Test concurrent worksheet processing."""
        # Add more worksheets
        sample_workbook.create_sheet("Sheet2")
        sample_workbook.create_sheet("Sheet3")
        
        # Add data to new sheets
        sample_workbook["Sheet2"]['A1'] = "Data1"
        sample_workbook["Sheet3"]['A1'] = "Data2"
        
        processed_worksheets = []
        
        async def mock_process_func(worksheet, cells):
            """Mock processing function."""
            processed_worksheets.append(worksheet.title)
            return len(cells)
        
        # Process worksheets concurrently
        total = await excel_processor.process_worksheets_concurrently(
            sample_workbook,
            mock_process_func,
            max_concurrency=2
        )
        
        # Verify all worksheets were processed
        assert len(processed_worksheets) == 3
        assert "Test Sheet" in processed_worksheets
        assert "Sheet2" in processed_worksheets
        assert "Sheet3" in processed_worksheets
        
        # Verify total count
        assert total > 0
    
    @pytest.mark.asyncio
    async def test_process_worksheets_with_progress_callback(self, excel_processor, sample_workbook):
        """Test worksheet processing with progress callback."""
        progress_updates = []
        
        async def progress_callback(progress: WorksheetProgress):
            """Capture progress updates."""
            progress_updates.append(progress)
        
        async def mock_process_func(worksheet, cells):
            """Mock processing function."""
            return len(cells)
        
        # Process with progress callback
        await excel_processor.process_worksheets_concurrently(
            sample_workbook,
            mock_process_func,
            progress_callback=progress_callback
        )
        
        # Verify progress updates were received
        assert len(progress_updates) > 0
        
        # Verify progress structure
        for progress in progress_updates:
            assert hasattr(progress, 'worksheet_name')
            assert hasattr(progress, 'cells_total')
            assert hasattr(progress, 'cells_processed')
            assert hasattr(progress, 'cells_translated')
    
    @pytest.mark.asyncio
    async def test_concurrent_processing_respects_limit(self, excel_processor):
        """Test that concurrent processing respects the concurrency limit."""
        # Create a workbook with many worksheets
        wb = Workbook()
        for i in range(15):
            ws = wb.create_sheet(f"Sheet{i}")
            ws['A1'] = f"Data{i}"
        
        # Remove default sheet
        wb.remove(wb['Sheet'])
        
        active_tasks = []
        max_concurrent = 0
        
        async def track_concurrency(worksheet, cells):
            """Track concurrent execution."""
            nonlocal max_concurrent
            active_tasks.append(worksheet.title)
            max_concurrent = max(max_concurrent, len(active_tasks))
            
            # Simulate some work
            await asyncio.sleep(0.01)
            
            active_tasks.remove(worksheet.title)
            return len(cells)
        
        # Process with limit of 5
        await excel_processor.process_worksheets_concurrently(
            wb,
            track_concurrency,
            max_concurrency=5
        )
        
        # Verify we never exceeded the limit
        assert max_concurrent <= 5


class TestCellData:
    """Tests for CellData dataclass."""
    
    def test_cell_data_creation(self):
        """Test creating a CellData instance."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        cell.value = "Test"
        
        cell_data = CellData(
            worksheet_name="Sheet1",
            row=1,
            column=1,
            value="Test",
            has_formula=False,
            cell=cell
        )
        
        assert cell_data.worksheet_name == "Sheet1"
        assert cell_data.row == 1
        assert cell_data.column == 1
        assert cell_data.value == "Test"
        assert cell_data.has_formula is False
        assert cell_data.cell == cell


class TestWorksheetProgress:
    """Tests for WorksheetProgress dataclass."""
    
    def test_worksheet_progress_creation(self):
        """Test creating a WorksheetProgress instance."""
        progress = WorksheetProgress(
            worksheet_name="Sheet1",
            cells_total=100,
            cells_processed=50,
            cells_translated=25
        )
        
        assert progress.worksheet_name == "Sheet1"
        assert progress.cells_total == 100
        assert progress.cells_processed == 50
        assert progress.cells_translated == 25
