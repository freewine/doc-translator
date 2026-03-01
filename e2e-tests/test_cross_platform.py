"""
Cross-platform API compatibility tests for Doc Translation System.

Tests API behavior, file handling, and platform-specific considerations
for the web-based translation system.
"""

import json
import os
import platform
import tempfile
import shutil
from pathlib import Path
from typing import Dict, List

import pytest
import requests
from openpyxl import Workbook

# Load test configuration
CONFIG_PATH = Path(__file__).parent / "e2e_config.json"
with open(CONFIG_PATH) as f:
    CONFIG = json.load(f)

BACKEND_URL = os.getenv("BACKEND_URL", CONFIG["test_environment"]["backend_url"])
GRAPHQL_ENDPOINT = os.getenv("GRAPHQL_ENDPOINT", CONFIG["test_environment"]["graphql_endpoint"])
HEALTH_ENDPOINT = CONFIG["test_environment"]["health_endpoint"]
UPLOAD_ENDPOINT = CONFIG["test_environment"]["upload_endpoint"]

TEST_USERNAME = CONFIG["test_credentials"]["username"]
TEST_PASSWORD = CONFIG["test_credentials"]["password"]


def authenticate() -> str:
    """Authenticate and return JWT token."""
    mutation = CONFIG["graphql_operations"]["login"]
    
    response = requests.post(
        GRAPHQL_ENDPOINT,
        json={
            "query": mutation,
            "variables": {"username": TEST_USERNAME, "password": TEST_PASSWORD}
        },
        headers={"Content-Type": "application/json"}
    )
    
    data = response.json()
    return data["data"]["login"]["token"]


def check_backend_running() -> bool:
    """Check if backend service is running."""
    try:
        response = requests.get(HEALTH_ENDPOINT, timeout=5)
        return response.status_code == 200
    except requests.RequestException:
        return False


class TestPlatformDetection:
    """Test platform detection and compatibility."""
    
    def test_current_platform_detected(self):
        """Test that current platform is correctly detected."""
        current_platform = platform.system()
        assert current_platform in ['Windows', 'Darwin', 'Linux']
        print(f"✓ Platform detected: {current_platform}")
    
    def test_path_separator_matches_platform(self):
        """Test that Path uses correct separator for platform."""
        test_path = Path('folder') / 'file.txt'
        path_str = str(test_path)
        
        if platform.system() == 'Windows':
            assert '\\' in path_str or '/' in path_str
        else:
            assert '/' in path_str
        
        print(f"✓ Path separator correct for {platform.system()}")


class TestFileHandling:
    """Test file handling across platforms."""
    
    @pytest.fixture
    def temp_dir(self):
        """Create a temporary directory for testing."""
        temp_path = tempfile.mkdtemp()
        yield temp_path
        shutil.rmtree(temp_path, ignore_errors=True)
    
    def test_create_document_file(self, temp_dir):
        """Test creating document files works on current platform."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "你好世界"
        ws['B1'] = "测试"
        
        file_path = Path(temp_dir) / "test.xlsx"
        wb.save(file_path)
        
        assert file_path.exists()
        assert file_path.is_file()
        assert file_path.stat().st_size > 0
        print(f"✓ Document file created: {file_path}")
    
    def test_path_normalization(self, temp_dir):
        """Test that paths are normalized correctly."""
        test_file = Path(temp_dir) / "test.xlsx"
        
        wb = Workbook()
        wb.active['A1'] = "test"
        wb.save(test_file)
        
        # Verify path is absolute and normalized
        assert test_file.is_absolute() or Path(temp_dir).is_absolute()
        assert test_file.exists()
        print(f"✓ Path normalization works correctly")
    
    def test_unicode_filename_handling(self, temp_dir):
        """Test handling of Unicode characters in filenames."""
        # Create file with Chinese characters in name
        unicode_filename = "测试文件_中文.xlsx"
        file_path = Path(temp_dir) / unicode_filename
        
        wb = Workbook()
        wb.active['A1'] = "内容"
        wb.save(file_path)
        
        assert file_path.exists()
        assert file_path.name == unicode_filename
        print(f"✓ Unicode filename handled: {unicode_filename}")
    
    def test_file_read_write_binary(self, temp_dir):
        """Test binary file read/write operations."""
        test_file = Path(temp_dir) / "test.xlsx"
        
        # Create and save
        wb = Workbook()
        wb.active['A1'] = "test content"
        wb.save(test_file)
        
        # Read binary
        content = test_file.read_bytes()
        assert content.startswith(b'PK')  # ZIP/XLSX signature
        
        # Write to new file
        new_file = Path(temp_dir) / "copy.xlsx"
        new_file.write_bytes(content)
        
        assert new_file.exists()
        assert new_file.stat().st_size == test_file.stat().st_size
        print(f"✓ Binary file operations work correctly")


@pytest.mark.skipif(not check_backend_running(), reason="Backend not running")
class TestAPICompatibility:
    """Test API compatibility across platforms."""
    
    @pytest.fixture
    def auth_token(self):
        """Get authentication token."""
        return authenticate()
    
    @pytest.fixture
    def temp_dir(self):
        """Create a temporary directory for testing."""
        temp_path = tempfile.mkdtemp()
        yield temp_path
        shutil.rmtree(temp_path, ignore_errors=True)
    
    def test_health_endpoint(self):
        """Test health endpoint responds correctly."""
        response = requests.get(HEALTH_ENDPOINT, timeout=5)
        
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "healthy"
        print(f"✓ Health endpoint: {data['status']}")
    
    def test_graphql_endpoint(self, auth_token):
        """Test GraphQL endpoint responds correctly."""
        query = "query { me { username } }"
        
        response = requests.post(
            GRAPHQL_ENDPOINT,
            json={"query": query},
            headers={
                "Authorization": f"Bearer {auth_token}",
                "Content-Type": "application/json"
            }
        )
        
        assert response.status_code == 200
        data = response.json()
        assert "errors" not in data
        assert data["data"]["me"]["username"] == TEST_USERNAME
        print(f"✓ GraphQL endpoint working")
    
    def test_file_upload_with_unicode_name(self, auth_token, temp_dir):
        """Test file upload with Unicode filename."""
        # Create test file with Unicode name
        unicode_filename = "测试文件_中文.xlsx"
        file_path = Path(temp_dir) / unicode_filename
        
        wb = Workbook()
        wb.active['A1'] = "你好世界"
        wb.save(file_path)
        
        # Upload file
        with open(file_path, 'rb') as f:
            files = {
                'file': (
                    unicode_filename,
                    f,
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            }
            response = requests.post(
                UPLOAD_ENDPOINT,
                files=files,
                headers={"Authorization": f"Bearer {auth_token}"}
            )
        
        assert response.status_code == 200
        data = response.json()
        assert "id" in data
        assert data["filename"] == unicode_filename
        print(f"✓ Unicode filename upload: {unicode_filename}")
    
    def test_file_upload_various_sizes(self, auth_token, temp_dir):
        """Test file upload with various file sizes."""
        sizes = [
            ("small", 10),      # 10 cells
            ("medium", 100),    # 100 cells
            ("large", 1000),    # 1000 cells
        ]
        
        for size_name, cell_count in sizes:
            # Create test file
            file_path = Path(temp_dir) / f"test_{size_name}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            for i in range(cell_count):
                row = (i // 10) + 1
                col = (i % 10) + 1
                ws.cell(row=row, column=col, value=f"测试内容{i}")
            wb.save(file_path)
            
            # Upload file
            with open(file_path, 'rb') as f:
                files = {
                    'file': (
                        file_path.name,
                        f,
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                }
                response = requests.post(
                    UPLOAD_ENDPOINT,
                    files=files,
                    headers={"Authorization": f"Bearer {auth_token}"}
                )
            
            assert response.status_code == 200
            data = response.json()
            print(f"✓ {size_name.capitalize()} file upload: {data['size']} bytes")
    
    def test_concurrent_requests(self, auth_token):
        """Test handling of concurrent API requests."""
        import concurrent.futures
        
        query = CONFIG["graphql_operations"]["get_language_pairs"]
        
        def make_request():
            response = requests.post(
                GRAPHQL_ENDPOINT,
                json={"query": query},
                headers={
                    "Authorization": f"Bearer {auth_token}",
                    "Content-Type": "application/json"
                }
            )
            return response.status_code == 200
        
        # Make 5 concurrent requests
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            futures = [executor.submit(make_request) for _ in range(5)]
            results = [f.result() for f in concurrent.futures.as_completed(futures)]
        
        assert all(results), "Some concurrent requests failed"
        print(f"✓ Concurrent requests: {len(results)} successful")


class TestMacOSSpecific:
    """Tests specific to macOS platform."""
    
    @pytest.mark.skipif(platform.system() != 'Darwin', reason="macOS-specific test")
    def test_macos_forward_slash_paths(self):
        """Test that macOS uses forward slashes."""
        test_path = Path('/Users/test/file.xlsx')
        assert '/' in str(test_path)
        assert '\\' not in str(test_path)
        print("✓ macOS forward slash paths")
    
    @pytest.mark.skipif(platform.system() != 'Darwin', reason="macOS-specific test")
    def test_macos_home_directory(self):
        """Test that home directory is accessible on macOS."""
        home = Path.home()
        assert home.exists()
        assert str(home).startswith('/Users/') or str(home).startswith('/home/')
        print(f"✓ macOS home directory: {home}")


class TestWindowsSpecific:
    """Tests specific to Windows platform."""
    
    @pytest.mark.skipif(platform.system() != 'Windows', reason="Windows-specific test")
    def test_windows_path_handling(self):
        """Test that Windows can handle various path formats."""
        # Windows should handle both forward and back slashes
        test_path = Path('C:/Users/test/file.xlsx')
        assert test_path.parts[0] == 'C:\\'
        print("✓ Windows path handling")
    
    @pytest.mark.skipif(platform.system() != 'Windows', reason="Windows-specific test")
    def test_windows_drive_letters(self):
        """Test that Windows drive letters are handled correctly."""
        test_path = Path('C:/test/file.xlsx')
        assert test_path.drive == 'C:'
        print("✓ Windows drive letters")
    
    @pytest.mark.skipif(platform.system() != 'Windows', reason="Windows-specific test")
    def test_windows_home_directory(self):
        """Test that home directory is accessible on Windows."""
        home = Path.home()
        assert home.exists()
        assert home.drive  # Should have a drive letter
        print(f"✓ Windows home directory: {home}")


class TestLinuxSpecific:
    """Tests specific to Linux platform."""
    
    @pytest.mark.skipif(platform.system() != 'Linux', reason="Linux-specific test")
    def test_linux_forward_slash_paths(self):
        """Test that Linux uses forward slashes."""
        test_path = Path('/home/test/file.xlsx')
        assert '/' in str(test_path)
        assert '\\' not in str(test_path)
        print("✓ Linux forward slash paths")
    
    @pytest.mark.skipif(platform.system() != 'Linux', reason="Linux-specific test")
    def test_linux_home_directory(self):
        """Test that home directory is accessible on Linux."""
        home = Path.home()
        assert home.exists()
        assert str(home).startswith('/home/') or str(home) == '/root'
        print(f"✓ Linux home directory: {home}")


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
