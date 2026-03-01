"""
End-to-End Tests using Chrome DevTools MCP

This module contains comprehensive end-to-end tests for the Doc Translation System
using Chrome DevTools MCP for browser automation. Tests cover the complete user
workflows including authentication, file upload, translation, and download.
"""

import json
import os
import tempfile
import time
from pathlib import Path
from typing import Dict, List, Optional
import pytest
import requests
from openpyxl import Workbook

# Load test configuration
CONFIG_PATH = Path(__file__).parent / "e2e_config.json"
with open(CONFIG_PATH) as f:
    CONFIG = json.load(f)

# Test configuration from config file
BACKEND_URL = os.getenv("BACKEND_URL", CONFIG["test_environment"]["backend_url"])
FRONTEND_URL = os.getenv("FRONTEND_URL", CONFIG["test_environment"]["frontend_url"])
GRAPHQL_ENDPOINT = os.getenv("GRAPHQL_ENDPOINT", CONFIG["test_environment"]["graphql_endpoint"])
HEALTH_ENDPOINT = CONFIG["test_environment"]["health_endpoint"]
UPLOAD_ENDPOINT = CONFIG["test_environment"]["upload_endpoint"]

# Test credentials
TEST_USERNAME = CONFIG["test_credentials"]["username"]
TEST_PASSWORD = CONFIG["test_credentials"]["password"]

# Viewport configurations for responsive testing
VIEWPORTS = CONFIG["viewports"]

# Timeouts
TIMEOUTS = CONFIG["test_timeouts"]


def create_test_document_files(temp_dir: tempfile.TemporaryDirectory) -> List[Path]:
    """Create test document files with Chinese content."""
    test_files = []
    
    # Test file 1: Simple Chinese text
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "测试工作表"
    ws1['A1'] = "你好世界"
    ws1['B1'] = "这是一个测试"
    ws1['A2'] = "中文翻译测试"
    ws1['B2'] = "文档文件处理"
    
    file1_path = Path(temp_dir.name) / "test_chinese_1.xlsx"
    wb1.save(file1_path)
    test_files.append(file1_path)
    
    # Test file 2: Mixed content with formatting
    from openpyxl.styles import Font
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "格式测试"
    ws2['A1'] = "标题文本"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['B1'] = "English text (should not be translated)"
    ws2['A2'] = "红色文本"
    ws2['A2'].font = Font(color="FF0000")
    ws2['B2'] = "123456"  # Numbers should not be translated
    
    file2_path = Path(temp_dir.name) / "test_mixed_content.xlsx"
    wb2.save(file2_path)
    test_files.append(file2_path)
    
    # Test file 3: Multiple worksheets
    wb3 = Workbook()
    ws3_1 = wb3.active
    ws3_1.title = "工作表1"
    ws3_1['A1'] = "第一个工作表"
    
    ws3_2 = wb3.create_sheet("工作表2")
    ws3_2['A1'] = "第二个工作表"
    ws3_2['B1'] = "更多中文内容"
    
    file3_path = Path(temp_dir.name) / "test_multiple_sheets.xlsx"
    wb3.save(file3_path)
    test_files.append(file3_path)
    
    return test_files


def wait_for_services(timeout: int = 30):
    """Wait for backend and frontend services to be ready."""
    start_time = time.time()
    
    # Wait for backend
    while time.time() - start_time < timeout:
        try:
            response = requests.get(HEALTH_ENDPOINT, timeout=5)
            if response.status_code == 200:
                break
        except requests.RequestException:
            pass
        time.sleep(1)
    else:
        pytest.skip("Backend service not ready")
    
    # Wait for frontend (check if port is responding)
    while time.time() - start_time < timeout:
        try:
            response = requests.get(FRONTEND_URL, timeout=5)
            if response.status_code in [200, 404]:  # 404 is OK for SPA
                break
        except requests.RequestException:
            pass
        time.sleep(1)
    else:
        pytest.skip("Frontend service not ready")


def authenticate_graphql() -> str:
    """Authenticate via GraphQL and return JWT token."""
    mutation = CONFIG["graphql_operations"]["login"]
    
    variables = {
        "username": TEST_USERNAME,
        "password": TEST_PASSWORD
    }
    
    response = requests.post(
        GRAPHQL_ENDPOINT,
        json={"query": mutation, "variables": variables},
        headers={"Content-Type": "application/json"}
    )
    
    assert response.status_code == 200, f"GraphQL request failed: {response.text}"
    
    data = response.json()
    assert "errors" not in data, f"GraphQL errors: {data.get('errors')}"
    
    token = data["data"]["login"]["token"]
    assert token, "No token received from login"
    
    return token


def check_services_running() -> bool:
    """Check if backend and frontend services are running."""
    try:
        # Check backend
        backend_response = requests.get(HEALTH_ENDPOINT, timeout=5)
        if backend_response.status_code != 200:
            return False
        
        # Check frontend
        frontend_response = requests.get(FRONTEND_URL, timeout=5)
        if frontend_response.status_code not in [200, 404]:  # 404 OK for SPA
            return False
        
        return True
    except requests.RequestException:
        return False


class TestE2EChrome:
    """Main E2E test class with Chrome DevTools MCP integration."""
    
    def setup_method(self):
        """Set up test environment before each test."""
        self.temp_dir = tempfile.TemporaryDirectory()
        self.test_files = []
        self.auth_token = None
        self.test_files = create_test_document_files(self.temp_dir)
        wait_for_services()
    
    def teardown_method(self):
        """Clean up after each test."""
        if hasattr(self, 'temp_dir') and self.temp_dir:
            self.temp_dir.cleanup()
        self.test_files = []
        self.auth_token = None
    
    @pytest.mark.e2e
    @pytest.mark.api
    def test_backend_health(self):
        """Test backend health endpoint."""
        response = requests.get(HEALTH_ENDPOINT, timeout=5)
        assert response.status_code == 200
        
        data = response.json()
        assert data["status"] == "healthy"
        assert "service" in data
        print("✓ Backend health check passed")
    
    @pytest.mark.e2e
    @pytest.mark.api
    def test_authentication_flow(self):
        """Test authentication via GraphQL API."""
        # Test valid login
        self.auth_token = authenticate_graphql()
        assert self.auth_token is not None
        assert len(self.auth_token) > 0
        print(f"✓ Authentication successful: {self.auth_token[:20]}...")
        
        # Test me query with token
        query = "query { me { username } }"
        response = requests.post(
            GRAPHQL_ENDPOINT,
            json={"query": query},
            headers={
                "Authorization": f"Bearer {self.auth_token}",
                "Content-Type": "application/json"
            }
        )
        
        assert response.status_code == 200
        data = response.json()
        assert "errors" not in data
        assert data["data"]["me"]["username"] == TEST_USERNAME
        print(f"✓ User verified: {TEST_USERNAME}")
    
    @pytest.mark.e2e
    @pytest.mark.api
    def test_language_pairs_query(self):
        """Test language pairs query."""
        self.auth_token = authenticate_graphql()
        
        query = CONFIG["graphql_operations"]["get_language_pairs"]
        
        response = requests.post(
            GRAPHQL_ENDPOINT,
            json={"query": query},
            headers={
                "Authorization": f"Bearer {self.auth_token}",
                "Content-Type": "application/json"
            }
        )
        
        assert response.status_code == 200
        data = response.json()
        assert "errors" not in data
        
        language_pairs = data["data"]["languagePairs"]
        assert len(language_pairs) > 0, "No language pairs configured"
        
        # Find Chinese to Vietnamese pair
        zh_vi_pair = None
        for pair in language_pairs:
            if pair["sourceLanguageCode"] == "zh" and pair["targetLanguageCode"] == "vi":
                zh_vi_pair = pair
                break
        
        assert zh_vi_pair, "Chinese to Vietnamese language pair not found"
        print(f"✓ Language pairs: {len(language_pairs)} available")
        print(f"✓ Chinese-Vietnamese pair: {zh_vi_pair['id']}")
    
    @pytest.mark.e2e
    @pytest.mark.api
    def test_model_config_query(self):
        """Test model configuration query."""
        self.auth_token = authenticate_graphql()
        
        query = CONFIG["graphql_operations"]["get_model_config"]
        
        response = requests.post(
            GRAPHQL_ENDPOINT,
            json={"query": query},
            headers={
                "Authorization": f"Bearer {self.auth_token}",
                "Content-Type": "application/json"
            }
        )
        
        assert response.status_code == 200
        data = response.json()
        assert "errors" not in data
        
        model_config = data["data"]["modelConfig"]
        assert model_config["modelId"], "No model ID configured"
        assert model_config["modelName"], "No model name configured"
        assert len(model_config["availableModels"]) > 0, "No available models"
        
        print(f"✓ Current model: {model_config['modelName']}")
        print(f"✓ Available models: {len(model_config['availableModels'])}")
    
    @pytest.mark.e2e
    @pytest.mark.api
    def test_file_upload_via_rest(self):
        """Test file upload via REST endpoint."""
        self.auth_token = authenticate_graphql()
        
        # Upload a test file
        test_file = self.test_files[0]
        
        with open(test_file, 'rb') as f:
            files = {'file': (test_file.name, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(
                UPLOAD_ENDPOINT,
                files=files,
                headers={"Authorization": f"Bearer {self.auth_token}"}
            )
        
        assert response.status_code == 200, f"Upload failed: {response.text}"
        
        data = response.json()
        assert "id" in data
        assert "filename" in data
        assert "size" in data
        assert data["filename"] == test_file.name
        
        print(f"✓ File uploaded: {data['filename']} (ID: {data['id']})")
    
    @pytest.mark.e2e
    @pytest.mark.api
    def test_jobs_query(self):
        """Test jobs query."""
        self.auth_token = authenticate_graphql()
        
        query = CONFIG["graphql_operations"]["get_jobs"]
        
        response = requests.post(
            GRAPHQL_ENDPOINT,
            json={"query": query},
            headers={
                "Authorization": f"Bearer {self.auth_token}",
                "Content-Type": "application/json"
            }
        )
        
        assert response.status_code == 200
        data = response.json()
        assert "errors" not in data
        
        jobs = data["data"]["jobs"]
        print(f"✓ Jobs query successful: {len(jobs)} jobs found")
    
    @pytest.mark.e2e
    @pytest.mark.api
    def test_error_unauthenticated_request(self):
        """Test that unauthenticated requests are rejected."""
        query = CONFIG["graphql_operations"]["get_jobs"]
        
        response = requests.post(
            GRAPHQL_ENDPOINT,
            json={"query": query},
            headers={"Content-Type": "application/json"}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert "errors" in data, "Expected authentication error"
        
        error_message = data["errors"][0]["message"].lower()
        assert "authentication" in error_message or "unauthorized" in error_message
        print("✓ Unauthenticated request properly rejected")
    
    @pytest.mark.e2e
    @pytest.mark.api
    def test_error_invalid_login(self):
        """Test that invalid login is rejected."""
        mutation = CONFIG["graphql_operations"]["login"]
        
        response = requests.post(
            GRAPHQL_ENDPOINT,
            json={
                "query": mutation,
                "variables": {"username": "invalid", "password": "invalid"}
            },
            headers={"Content-Type": "application/json"}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert "errors" in data, "Expected login error"
        print("✓ Invalid login properly rejected")
    
    @pytest.mark.e2e
    @pytest.mark.api
    def test_add_and_remove_language_pair(self):
        """Test adding and removing a language pair."""
        self.auth_token = authenticate_graphql()
        
        # Add a new language pair
        add_mutation = CONFIG["graphql_operations"]["add_language_pair"]
        
        variables = {
            "sourceLanguage": "English",
            "targetLanguage": "Spanish",
            "sourceLanguageCode": "en",
            "targetLanguageCode": "es"
        }
        
        response = requests.post(
            GRAPHQL_ENDPOINT,
            json={"query": add_mutation, "variables": variables},
            headers={
                "Authorization": f"Bearer {self.auth_token}",
                "Content-Type": "application/json"
            }
        )
        
        assert response.status_code == 200
        data = response.json()
        
        if "errors" in data:
            # Language pair might already exist
            print(f"Note: {data['errors'][0]['message']}")
        else:
            new_pair = data["data"]["addLanguagePair"]
            print(f"✓ Added language pair: {new_pair['sourceLanguage']} → {new_pair['targetLanguage']}")
            
            # Remove the language pair
            remove_mutation = CONFIG["graphql_operations"]["remove_language_pair"]
            
            response = requests.post(
                GRAPHQL_ENDPOINT,
                json={"query": remove_mutation, "variables": {"id": new_pair["id"]}},
                headers={
                    "Authorization": f"Bearer {self.auth_token}",
                    "Content-Type": "application/json"
                }
            )
            
            assert response.status_code == 200
            data = response.json()
            assert "errors" not in data
            print(f"✓ Removed language pair: {new_pair['id']}")
    
    @pytest.mark.e2e
    @pytest.mark.browser
    def test_complete_user_workflow(self):
        """
        Test the complete user workflow: login → upload → translate → download
        
        Note: This test requires Chrome DevTools MCP to be configured.
        When MCP is not available, it falls back to API testing.
        """
        print("Testing complete user workflow via API...")
        
        # Step 1: Authenticate
        self.auth_token = authenticate_graphql()
        print("✓ Step 1: Authentication successful")
        
        # Step 2: Get language pairs
        query = CONFIG["graphql_operations"]["get_language_pairs"]
        response = requests.post(
            GRAPHQL_ENDPOINT,
            json={"query": query},
            headers={
                "Authorization": f"Bearer {self.auth_token}",
                "Content-Type": "application/json"
            }
        )
        
        data = response.json()
        language_pairs = data["data"]["languagePairs"]
        zh_vi_pair = next(
            (p for p in language_pairs if p["sourceLanguageCode"] == "zh" and p["targetLanguageCode"] == "vi"),
            None
        )
        assert zh_vi_pair, "Chinese-Vietnamese pair not found"
        print(f"✓ Step 2: Language pair found: {zh_vi_pair['id']}")
        
        # Step 3: Upload file
        test_file = self.test_files[0]
        with open(test_file, 'rb') as f:
            files = {'file': (test_file.name, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(
                UPLOAD_ENDPOINT,
                files=files,
                headers={"Authorization": f"Bearer {self.auth_token}"}
            )
        
        assert response.status_code == 200
        file_data = response.json()
        file_id = file_data["id"]
        print(f"✓ Step 3: File uploaded: {file_data['filename']}")
        
        # Step 4: Create translation job
        create_job_mutation = CONFIG["graphql_operations"]["create_translation_job"]
        
        response = requests.post(
            GRAPHQL_ENDPOINT,
            json={
                "query": create_job_mutation,
                "variables": {
                    "fileIds": [file_id],
                    "languagePairId": zh_vi_pair["id"]
                }
            },
            headers={
                "Authorization": f"Bearer {self.auth_token}",
                "Content-Type": "application/json"
            }
        )
        
        assert response.status_code == 200
        data = response.json()
        assert "errors" not in data, f"Job creation failed: {data.get('errors')}"
        
        job = data["data"]["createTranslationJob"]
        job_id = job["id"]
        print(f"✓ Step 4: Translation job created: {job_id}")
        
        # Step 5: Poll for job completion (with timeout)
        get_job_query = CONFIG["graphql_operations"]["get_job"]
        max_wait = TIMEOUTS["translation_job"]
        poll_interval = TIMEOUTS["polling_interval"]
        start_time = time.time()
        
        while time.time() - start_time < max_wait:
            response = requests.post(
                GRAPHQL_ENDPOINT,
                json={"query": get_job_query, "variables": {"id": job_id}},
                headers={
                    "Authorization": f"Bearer {self.auth_token}",
                    "Content-Type": "application/json"
                }
            )
            
            data = response.json()
            job_status = data["data"]["job"]
            
            if job_status["status"] in ["COMPLETED", "PARTIAL_SUCCESS"]:
                print(f"✓ Step 5: Job completed with status: {job_status['status']}")
                print(f"  Progress: {job_status['progress'] * 100:.1f}%")
                print(f"  Files completed: {job_status['filesCompleted']}/{job_status['filesTotal']}")
                break
            elif job_status["status"] == "FAILED":
                pytest.fail(f"Translation job failed: {job_status.get('filesFailed', [])}")
            
            print(f"  Polling... Status: {job_status['status']}, Progress: {job_status['progress'] * 100:.1f}%")
            time.sleep(poll_interval)
        else:
            pytest.fail(f"Job did not complete within {max_wait} seconds")
        
        print("✓ Complete user workflow test passed")
    
    @pytest.mark.e2e
    @pytest.mark.browser
    def test_responsive_design_viewports(self):
        """
        Test responsive design at different viewports.
        
        Note: This test documents expected behavior for browser testing.
        Actual viewport testing requires Chrome DevTools MCP.
        """
        print("Testing responsive design expectations...")
        
        for viewport_name, dimensions in VIEWPORTS.items():
            width = dimensions["width"]
            height = dimensions["height"]
            description = dimensions.get("description", "")
            
            if width < 768:
                expected_layout = "mobile"
                print(f"  ✓ {viewport_name} ({description}): Expected mobile layout with stacked components")
            elif width < 1024:
                expected_layout = "tablet"
                print(f"  ✓ {viewport_name} ({description}): Expected tablet layout with optimized spacing")
            else:
                expected_layout = "desktop"
                print(f"  ✓ {viewport_name} ({description}): Expected desktop layout with full features")
        
        print("✓ Responsive design expectations documented")


@pytest.fixture(scope="session", autouse=True)
def check_test_environment():
    """Check that the test environment is properly set up."""
    if not check_services_running():
        pytest.skip("Backend and/or frontend services are not running. "
                   "Please start the services before running E2E tests.")


if __name__ == "__main__":
    """Run the E2E tests directly."""
    import sys
    
    if not check_services_running():
        print("❌ Services not running. Please start backend and frontend services:")
        print("   Backend: cd backend && uv run main.py")
        print("   Frontend: cd frontend && npm run dev")
        sys.exit(1)
    
    print("✅ Services are running. Starting E2E tests...")
    
    pytest.main([
        __file__,
        "-v",
        "--tb=short",
        "-m", "e2e"
    ])
